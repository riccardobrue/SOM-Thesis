from openpyxl import load_workbook
import pandas as pd
import numpy as np
import os, os.path

# parameters
typology = "EQUAL"  # UNEQUAL
selected_height = "50"
selected_width = "50"

# Basic path (where the data is stored)
basic_path = "C:\\Users\\Riccardo\\Google Drive\\University\\Double Degree - Middlesex\\Middlesex Teaching Material\\CSD4444 - Ralph Moseley\\Data\\"
# Excel structure folder
excel_folder_path = basic_path + "Real data\\data_science_simulation\\" + typology + "_BS_75_D4\\" + "AREA_H" + selected_height + "\\"
# Simulation folders
simulation_folder_path = basic_path + "Real data\\SimulationEQUAL\\"
# Output folder
output_folder_path = basic_path + "Manipulated data\\"
# input excel file
input_file = "AREA_" + selected_height + "x" + selected_width + ".xlsx"
# output csv (merged) file
output_merged_file = "merged_equal_" + selected_height + "x" + selected_width + ".csv"


def get_data(restore=False):
    if restore:
        # ----------------------------------------------------------
        # READ ALREADY EXTRACTED DATA FROM CSV
        # ----------------------------------------------------------
        dataframe = pd.read_csv(output_folder_path + output_merged_file)
    else:
        # ----------------------------------------------------------
        # MERGE SIMULATED RESULTS WITH EXCEL STRUCTURE AND CREATE CSV FILE
        # ----------------------------------------------------------
        excel_file = excel_folder_path + input_file

        list = os.listdir(simulation_folder_path)  # dir is your directory path
        number_files = len(list)

        selected_files = []

        for i in range(0, len(list)):
            current_file_name = list[i]

            width, height, hom_energy, hom_rate, perc_aggr, heterogeneity, transm_range = current_file_name.split("_")

            if width[1:] == selected_width and height[1:] == selected_height:
                selected_files.append(selected_files)
                print(len(selected_files), ") ", current_file_name)


        """
        
        wb = load_workbook(complete_file_path, data_only=True)  # , read_only=True)

        # selecting the first (or active) sheet
        ws = wb.active
        # or
        # first_sheet = wb.get_sheet_names()[0]
        # ws = wb.get_sheet_by_name(first_sheet)

        columns = np.array(['A', 'B', 'C', 'D', 'E', 'F'])  # is the first output column
        num_attributes = 5

        data_dictionary = {}
        max_rows = 0  # the first column defines the maximum number of rows
        for c in range(len(columns)):
            column = ws[columns[c]]
            column_header = str(column[0].value)
            data_dictionary.update({column_header: []})

            for x in range(1, len(column)):
                if column[x].value is None:
                    if c == 0:
                        max_rows = x
                        break
                    elif c > 0 and x < max_rows:
                        data_dictionary[column_header].append(column[x].value)
                    else:
                        break
                else:
                    data_dictionary[column_header].append(column[x].value)
        dataframe = pd.DataFrame(data=data_dictionary, columns=data_dictionary.keys())

        # print(dataframe)

        """
        """
        # ----------------------------------------------------------
        # MANIPULATING DATA
        # ----------------------------------------------------------
        dataframe['HEED FND'] = dataframe.apply(
            lambda row: row['R0'] * row['HOM ENERGY'],
            # lambda row: row['R0']*row['HOM ENERGY'] if np.isnan(row['c']) else row['c'],
            axis=1
        )
        # print(dataframe)
        """
        """
        # ----------------------------------------------------------
        # STORING DATA TO A SIMPLEST CSV
        # ----------------------------------------------------------
        dataframe.to_csv(manipulated_csv_name, index=False)
        
    # ----------------------------------------------------------
    # CONVERT DATAFRAME TO NP ARRAY
    # ----------------------------------------------------------
    column_names = list(dataframe.columns.values)
    input_cols = [column_names[0], column_names[1], column_names[2], column_names[3], column_names[4]]
    output_cols = [column_names[5]]

    inputs = dataframe.as_matrix(input_cols)
    efficiency_values = dataframe.as_matrix(output_cols)
    efficiency_values = np.squeeze(efficiency_values)
    efficiency_values = list(efficiency_values)

    return inputs, efficiency_values
"""


if __name__ == "__main__":
    get_data()
