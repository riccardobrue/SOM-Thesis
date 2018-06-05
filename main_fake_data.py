from matplotlib import pyplot as plt
from SOM import SOM
import numpy as np
from openpyxl import load_workbook
import pandas as pd

# ----------------------------------------------------------
# LOAD DATA FROM FILE
# ----------------------------------------------------------
data_path = "C:\\Users\\Riccardo\\Google Drive\\University\\Double Degree - Middlesex\\Middlesex Teaching Material\\CSD4444 - Ralph Moseley\\Data\\Temporary fake data\\"
data_file_name = "AREA_50x50.xlsx"
complete_file_path = data_path + data_file_name

wb = load_workbook(complete_file_path)  # , read_only=True)

# selecting the first (or active) sheet
ws = wb.active
# or
# first_sheet = wb.get_sheet_names()[0]
# ws = wb.get_sheet_by_name(first_sheet)

columns = np.array(['A', 'B', 'C', 'D', 'E', 'F'])  # is the first output column

data_dictionary = {}
max_rows = 0  # the first column defines the maximum number of rows
for c in range(len(columns)):
    column = ws[columns[c]]
    column_header = str(column[0].value)
    data_dictionary.update({column_header: []})

    for x in range(1, len(column)):
        if column[x].value is None:
            if c == 0:
                max_rows = x
                break
            elif c > 0 and x < max_rows:
                data_dictionary[column_header].append(column[x].value)
            else:
                break
        else:
            data_dictionary[column_header].append(column[x].value)
dataframe = pd.DataFrame(data=data_dictionary, columns=data_dictionary.keys())

print(dataframe)
# ----------------------------------------------------------
# MANIPULATING DATA
# ----------------------------------------------------------


# ----------------------------------------------------------

# Training inputs for RGB colors (percentage of Red, Green and Blue)
inputs = np.array(
    [[0., 0., 0.],  # black
     [0., 0., 1.],  # blue
     [0., 0., 0.5],  # dark blue
     [0.125, 0.529, 1.0],  # sky blue
     [0.33, 0.4, 0.67],  # grey blue
     [0.6, 0.5, 1.0],  # lilac
     [0., 1., 0.],  # green
     [1., 0., 0.],  # red
     [0., 1., 1.],  # cyan
     [1., 0., 1.],  # violet
     [1., 1., 0.],  # yellow
     [1., 1., 1.],  # white
     [.33, .33, .33],  # dark grey
     [.5, .5, .5],  # medium grey
     [.66, .66, .66]  # light grey
     ])

efficiency_values = ['black', 'blue', 'darkblue', 'skyblue',
                     'greyblue', 'lilac', 'green', 'red',
                     'cyan', 'violet', 'yellow', 'white',
                     'darkgrey', 'mediumgrey', 'lightgrey']

# Train a 20x30 SOM (2D) with 400 iterations
# 20: SOM's grid height
# 30: SOM's grid width
# 3: input dimensionality (RGB values)
# 400: iterations
som = SOM(20, 30, 3, 100)

som.train(inputs, training_graph=False, labels=efficiency_values)  # training the SOM with the input data

# Get output grid
image_grid = som.get_centroids()

# Map colours to their closest neurons
mapped = som.map_vects(inputs)

print("Plotting final chart")
# Plot
plt.imshow(image_grid)
plt.title('Color SOM')
for i, m in enumerate(mapped):
    plt.text(m[1], m[0], efficiency_values[i], ha='center', va='center', bbox=dict(facecolor='white', alpha=0.5, lw=0))

plt.show()
