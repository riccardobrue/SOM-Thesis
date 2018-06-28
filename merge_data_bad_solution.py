from openpyxl import load_workbook
import pandas as pd
import numpy as np
import os, os.path

# parameters
typology = "EQUAL"  # UNEQUAL
selected_height = "50"
selected_width = "50"

# Basic path (where the data is stored)
basic_path = "C:\\Users\\Riccardo\\Google Drive\\University\\Double Degree - Middlesex\\Middlesex Teaching Material\\CSD4444 - Ralph Moseley\\Data\\"
# Excel structure folder
excel_folder_path = basic_path + "Real data\\data_science_simulation\\" + typology + "_BS_75_D4\\" + "AREA_H" + selected_height + "\\"
# Simulation folders
simulation_folder_path = basic_path + "Real data\\SimulationEQUAL\\"
# Output folder
output_folder_path = basic_path + "Manipulated data\\"
# input excel file
input_file = "AREA_" + selected_height + "x" + selected_width + ".xlsx"
# output csv (merged) file
output_merged_file_path = output_folder_path + "merged_equal_" + selected_height + "x" + selected_width + ".csv"


def get_data(restore=False):
    if restore:
        # ----------------------------------------------------------
        # READ ALREADY EXTRACTED DATA FROM CSV
        # ----------------------------------------------------------
        dataframe = pd.read_csv(output_merged_file_path)
    else:
        # ----------------------------------------------------------
        # MERGE SIMULATED RESULTS WITH EXCEL STRUCTURE AND CREATE CSV FILE
        # ----------------------------------------------------------
        # ----------------------------------------------------------
        # SELECT THE RELEVANT SIMULATION RESULT FILES
        # ----------------------------------------------------------
        list = os.listdir(simulation_folder_path)  # dir is your directory path
        selected_files = []

        for i in range(0, len(list)):
            current_file_name = list[i]

            width, height, hom_energy, hom_rate, perc_aggr, heterogeneity, transm_range_e = current_file_name.split("_")
            """
            width:              W...        [1:] 
            height:             H...        [1:] 
            hom_energy:         Henergy...  [7:] 
            hom_rate:           Hrate...    [5:]
            perc_aggr:          Agg...      [3:]
            heterogeneity:      Hperc...    [5:]
            transm_range_e:     R0...       [2:]
            """

            width = width[1:]
            height = height[1:]

            hom_energy = hom_energy[7:]
            if hom_energy == "0.75":
                hom_energy = 0.5
            elif hom_energy == "0.5":
                hom_energy = 0.3

            hom_rate = hom_rate[5:]

            perc_aggr = perc_aggr[3:]
            if perc_aggr == "100percentAggregation":
                perc_aggr = 1
            elif perc_aggr == "70percentAggregation":
                perc_aggr = 0.6
            elif perc_aggr == "40percentAggregation":
                perc_aggr = 0.3
            elif perc_aggr == "NoAggregation":
                perc_aggr = 0

            heterogeneity = heterogeneity[5:]
            if heterogeneity == "0.7":
                heterogeneity = 0.6
            elif heterogeneity == "0.4":
                heterogeneity = 0.3

            transm_range, extension = str(transm_range_e[2:]).split(".")

            if width == selected_width and height == selected_height:
                selected_files.append(selected_files)
                print(len(selected_files), ") ", current_file_name, " - ", hom_energy, ", ", hom_rate, ", ", perc_aggr,
                      ", ", heterogeneity, ", ", transm_range)

        # ----------------------------------------------------------
        # LOAD THE RELEVANT EXCEL FILE
        # ----------------------------------------------------------
        excel_file = excel_folder_path + input_file
        wb = load_workbook(excel_file, data_only=True)  # , read_only=True)

        # selecting the first (or active) sheet
        ws = wb.active
        # or
        # first_sheet = wb.get_sheet_names()[0]
        # ws = wb.get_sheet_by_name(first_sheet)

        columns = np.array(['A', 'B', 'C', 'D', 'E'])  # is the first output column
        num_attributes = 5

        data_dictionary = {}
        max_rows = 0  # the first column defines the maximum number of rows
        for c in range(len(columns)):
            column = ws[columns[c]]
            column_header = str(column[0].value)
            data_dictionary.update({column_header: []})

            for x in range(1, len(column)):
                if column[x].value is None:
                    if c == 0:
                        max_rows = x
                        break
                    elif c > 0 and x < max_rows:
                        data_dictionary[column_header].append(column[x].value)
                    else:
                        break
                else:
                    data_dictionary[column_header].append(column[x].value)
        dataframe = pd.DataFrame(data=data_dictionary, columns=data_dictionary.keys())

        # ----------------------------------------------------------
        # MERGE EXCEL STRUCTURE WITH SIMULATION DATA
        # ----------------------------------------------------------

        # add 4 empty columns for the protocol results (REECHD, HEED, ERHEED, FMUC)
        dataframe["REECHD"] = None
        dataframe["HEED"] = None
        dataframe["ERHEED"] = None
        dataframe["FMUC"] = None

        print(dataframe)

        for index, row in dataframe.iterrows():
            hom_energy = row['HOM ENERGY']
            hom_rate = row['HOM RATE ']
            perc_aggr = row['%AGGR']
            heterogeneity = row['HET']
            transm_range = row['R0']

            print(hom_energy, hom_rate, perc_aggr, heterogeneity, transm_range)

        """
        #print(dataframe)
        # ----------------------------------------------------------
        # MANIPULATING DATA
        # ----------------------------------------------------------
        dataframe['HEED FND'] = dataframe.apply(
            lambda row: row['R0'] * row['HOM ENERGY'],
            # lambda row: row['R0']*row['HOM ENERGY'] if np.isnan(row['c']) else row['c'],
            axis=1
        )
        """
        # print(dataframe)

        # ----------------------------------------------------------
        # STORING DATA TO A SIMPLEST CSV
        # ----------------------------------------------------------
        dataframe.to_csv(output_merged_file_path, index=False)


"""
    # ----------------------------------------------------------
    # CONVERT DATAFRAME TO NP ARRAY
    # ----------------------------------------------------------
    column_names = list(dataframe.columns.values)
    input_cols = [column_names[0], column_names[1], column_names[2], column_names[3], column_names[4]]
    output_cols = [column_names[5]]

    inputs = dataframe.as_matrix(input_cols)
    efficiency_values = dataframe.as_matrix(output_cols)
    efficiency_values = np.squeeze(efficiency_values)
    efficiency_values = list(efficiency_values)

    return inputs, efficiency_values
"""

if __name__ == "__main__":
    get_data()
